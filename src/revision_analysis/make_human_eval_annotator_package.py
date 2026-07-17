from __future__ import annotations

import csv
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def make_xlsx(src: Path, out: Path) -> None:
    rows = list(csv.DictReader(src.open(encoding="utf-8-sig", newline="")))
    wb = Workbook()
    ws = wb.active
    ws.title = "blind_eval"
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "item_id": 8,
        "pair": 10,
        "strategy": 10,
        "source_text": 45,
        "reference_translation": 45,
        "candidate_label": 12,
        "candidate_translation": 55,
        "adequacy_1_5": 14,
        "fluency_1_5": 12,
        "overall_preference_A_B_Tie": 24,
        "error_type": 18,
        "comments": 35,
    }
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    inst = wb.create_sheet("instructions")
    instructions = [
        "请填写 blind_eval 里的空白列。",
        "adequacy_1_5: 忠实度，1 最差，5 最好。",
        "fluency_1_5: 流畅度，1 最差，5 最好。",
        "overall_preference_A_B_Tie: 每个 item 只填 A、B 或 Tie。",
        "error_type: 可选，建议 incomplete / mistranslation / omission / addition / terminology / fluency / style / other。",
        "comments: 可选，简短说明判断原因。",
        "不要改 candidate_label、source_text、reference_translation、candidate_translation。",
    ]
    for line in instructions:
        inst.append([line])
    inst.column_dimensions["A"].width = 120
    for cell in inst["A"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out)


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    package_dir = root / "revision_analysis" / "human_eval_to_send_annotators"
    package_dir.mkdir(parents=True, exist_ok=True)
    src = package_dir / "qualitative_error_case_blind_sheet.csv"
    xlsx = package_dir / "qualitative_error_case_blind_sheet.xlsx"
    make_xlsx(src, xlsx)

    zip_path = root / "revision_analysis" / "human_eval_annotator_package.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in [
            package_dir / "qualitative_error_case_blind_sheet.csv",
            package_dir / "qualitative_error_case_blind_sheet.xlsx",
            package_dir / "HUMAN_EVAL_PROTOCOL_zh.md",
        ]:
            zf.write(path, arcname=path.name)
    print(zip_path)


if __name__ == "__main__":
    main()
